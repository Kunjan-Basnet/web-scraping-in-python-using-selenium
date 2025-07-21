import os
import time
import random
import pandas as pd
from random import randint, uniform
from seleniumbase import SB
from helping_functions import *
from seleniumbase.config import settings
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import openpyxl
from tqdm import tqdm
import logging
import pywhatkit as kit
from datetime import datetime
import ast
import sys
from selenium.common.exceptions import NoSuchElementException

logging.basicConfig(filename='example.log', level=logging.DEBUG)
phone_number = "123456789"
pc = "PC1"


links = [""]
paths = [""]
base_link = 'https://xyz.com/'
base_dir = os.path.dirname(os.path.abspath(__file__))
base_path = base_dir + r"\drawer rail"
print("Base Path: ", base_path)



sleep_time = 3
link_list = None
whole_page_load = True
header = False

table_and_rows=[]
tables = None
table_count = 0
count = 0
filename = ""

address=whats_my_address()

def send_notification(message = None):
    try:
        current_time = datetime.now()
        hour = current_time.hour
        minute = current_time.minute+1
        response = kit.sendwhatmsg(phone_number, message, hour, minute)
    except:
        pass

def remove_duplicates_and_sort(file_path, sheet_number, part_numbers_data):
    """
    Removes duplicate rows and sorts the specified sheet in an Excel file based on the provided list of part numbers.

    Args:
        file_path (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to process.
        part_numbers_data (list): A list of reference part numbers in the desired order.
    """

    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Get the specified worksheet
    sheet_name = f"Sheet{sheet_number}"
    worksheet = workbook[sheet_name]

    # Create a dictionary to map part numbers to rows
    part_number_to_row = {}

    # Iterate over the rows, skipping the header row
    for row in worksheet.iter_rows(min_row=2):
        part_number = str(row[0].value)  # Assuming the part number is in the first column
        if part_number in part_numbers_data:
            part_number_to_row[part_number] = tuple(cell.value for cell in row)

    # Remove duplicates by creating a list of unique rows in the order specified by part_numbers_data
    sorted_rows = [part_number_to_row[part] for part in part_numbers_data if part in part_number_to_row]

    # Clear the existing data but keep the header
    worksheet.delete_rows(2, worksheet.max_row - 1)

    # Append the sorted rows
    for row_values in sorted_rows:
        worksheet.append(row_values)

    # Save the workbook
    workbook.save(file_path)

def scroll_whole_page(driver):
    global tables, table_count
    body = driver.find_element(By.TAG_NAME, "body")
    # body.click()
    active_element = driver.switch_to.active_element
    is_tables_visible = check_visibility(driver, "//table")
    sleep(uniform(sleep_time, sleep_time+1))
    if is_tables_visible:
        tables = driver.find_elements(By.CSS_SELECTOR, ".hn")
        sleep(uniform(sleep_time, sleep_time+1))
        table_count = len(tables)
    else:
        send_notification(message = f'Could not load webpage in {pc}.')
        return False
    print("Scrolling Down to load webpage in cache....")
    for table_index in tqdm(range(table_count), desc="Processing", leave=True):
        try:
            tables = driver.find_elements(By.CSS_SELECTOR, ".hn")
            sleep(uniform(sleep_time, sleep_time+1))
            scroll_to_view(driver, tables[table_index])
            sleep(uniform(sleep_time, sleep_time+1))
        except:
            body.click()
            sleep(uniform(sleep_time, sleep_time+1))
            try:
                temp_div = driver.find_element(By.XPATH, "//div[@class='InLnOrdWebPartLayout_Overlay InLnOrdWebPart_RacingStripe']")
                div_id = temp_div.get_attribute('id').split("Stripe")[-1]
                link_element = driver.find_element(By.LINK_TEXT, div_id)
                link_element.click()
            except:
                pass
            sleep(uniform(sleep_time, sleep_time+1))
            active_element.send_keys(Keys.PAGE_DOWN)
            

def scrape_table_data(driver, table_number, remain, final_store_path, file_path):
    global header, sleep_time
    final_data = []
    is_tables_visible = check_visibility(driver, "//table")
    sleep(uniform(sleep_time, sleep_time+1))
    if is_tables_visible:
        tables = driver.find_elements(By.CSS_SELECTOR, ".hn")
    else:
        send_notification(message = f'Could not load webpage in {pc}.')
        return False
        
    current_table = tables[table_number-1]
    scroll_to_view(driver, current_table)
    sleep(uniform(sleep_time, sleep_time+1))
    try:
        anchor_tag = driver.find_element(By.XPATH, f"//tr//a[contains(text(), '{remain}')]")
        scroll_to_view(driver, anchor_tag)
        sleep(uniform(sleep_time, sleep_time+1))
        data = {}
        price_value=driver.find_element(By.XPATH,f"//td//a[contains(text(),'{remain}')]//parent::td/following-sibling::td").text
        if price_value.startswith("$") or price_value[0].isdigit():
                price_value = price_value.replace("$", "")
        else:
            return None

        
        
        try:
            print(f"***********Scraping data for part number: {remain}*************")
            data["part_number"] = remain
            data["brand"] = "xyz"
            data["retailer"] = "xyz"
            data["address"]= address
            data['pieces_per_package'] = "1"
            anchor_tag.click()
            current_time=datetime.now()
            data['extracted_date'] = current_time.timestamp()
            sleep(uniform(sleep_time, sleep_time+1))
            try:
                try:
                    choose_circumference = driver.find_element(By.XPATH,"//div[contains(@class,'kw')]//input[contains(@class,'ch')]")
                    choose_circumference.send_keys("100")
                except:
                    pass

                try:
                    choose_circumference=driver.find_element(By.XPATH,"//td[contains(@class,'mg ng nj SpecSrch_Value')]")
                    choose_circumference.click()
                except:
                    pass

                quantity_input = driver.find_element(By.XPATH, "//input[contains(@id,'qtyInp')]")
                quantity_input.send_keys("1")

                try:
                    choose_color=driver.find_element(By.XPATH,"//td[contains(@class,'me mr na ng SpecSrch_Value')]")
                    choose_color.click()
                except:
                    pass
                sleep_time += 1
                sleep(uniform(sleep_time, sleep_time+1))
            except:
                anchor_tag.click()
                sleep(uniform(sleep_time, sleep_time+1))
                
                try:
                    choose_circumference = driver.find_element(By.XPATH,"//div[contains(@class,'kw')]//input[contains(@class,'ch')]")
                    choose_circumference.send_keys("100")
                except:
                    pass

                try:
                    choose_circumference=driver.find_element(By.XPATH,"//td[contains(@class,'mg ng nj SpecSrch_Value')]")
                    choose_circumference.click()
                except:
                    pass

                quantity_input = driver.find_element(By.XPATH, "//input[contains(@id,'qtyInp')]")
                quantity_input.send_keys("1")


                try:
                    choose_color=driver.find_element(By.XPATH,"//td[contains(@class,'me mr na ng SpecSrch_Value')]")
                    choose_color.click()
                except:
                    pass
                sleep_time += 1
                sleep(uniform(sleep_time, sleep_time+1))
            sleep_time=1    
            add_to_order_btn = driver.find_element(By.XPATH, "//button[contains(@class,'add-to-order')]")
            add_to_order_btn.click()
            sleep(uniform(sleep_time, sleep_time+2))
            delivery_date_text = ""
            while delivery_date_text == "":
                delivery_date = driver.find_element(By.XPATH, "//div[contains(@class,'InLnOrdWebPartLayout_Main')]").text
                for t in ['Deliver']:
                    if t.lower() in delivery_date.lower():
                        delivery_date_text =separate_text(delivery_date) 
                        data[f"days_to_ship_1"] = delivery_date_text.strip()
                        

                        break
                    else:
                        sleep_time += 1
                        sleep(uniform(sleep_time, sleep_time+1))
                        continue
            anchor_tag.click()
            sleep(uniform(sleep_time, sleep_time+1))
            # if pack_value !=1:
            #     data["quantity_range_start_1"] =""
            #     data["quantity_range_end_1"] =""
            #     data["unit_price_1"] =""
            #     data["pieces_per_package_1"]=pack_value
            #     data["price_1"]=price_value
            #     data["unit_1"]="$"
            # else:
            
            data["quantity_range_start_1"] ="1"
            data["quantity_range_end_1"] = "0"
            data["price_1"] = price_value
            # data["pieces_per_package_1"]=""
            # data["price_1"]=""
            data["unit_1"]="$"
            final_data.append(data)
        except Exception as e:
            print(e)
            send_notification(message = f'Data Extraction error in {remain} in {pc}.')
            logging.error(f'Data Extraction error in {remain}.')
                
    except Exception as e:
        print("Some part numbers are not extracted.")
        print(e)
    finally:
        # for d in data.values():
        #     print(len(d))
        df = pd.DataFrame(final_data)
        df.dropna(how="all", inplace=True)
        

        # Load existing workbook if it exists
        wb = load_workbook(file_path, data_only=True)
        
        # Get the sheet to write to
        sheet = wb[f'Sheet{table_number}']
        # Append data as rows
        for row in df.to_numpy():  # Convert DataFrame to NumPy array for efficiency
            sheet.append(row.tolist())

        # Save the workbook
        wb.save(f'{final_store_path}/{count}_{filename}.xlsx')


def preparing_link_file(driver, url, final_store_path):
    try:
        driver.get(url)
        section_element_visibility = check_visibility(driver, "//section[contains(@class,'GroupPrsnttn')]")
        sleep(uniform(sleep_time, sleep_time+2))
        if section_element_visibility:
            link_hrefs = []
            section_elements = driver.find_elements(By.XPATH, "//section[contains(@class,'GroupPrsnttn')]")
            for section_element in section_elements:
                link_elements = section_element.find_elements(By.XPATH, ".//a")
                for link in link_elements:
                    link_hrefs.append(link.get_attribute('href'))
            df = pd.DataFrame({'Link Href': link_hrefs})
            df.to_csv(f"{final_store_path}/00_part_numbers_links.csv", index=False)
            print("Files with links created successfully.")
            print()
            sleep(uniform(sleep_time, sleep_time+2))
        else:
            send_notification(message = f'Could not load webpage in {pc}.')
            return False
    except  Exception as e:
        print(e)

            
def start_scraping(url, final_store_path, start_num=1):
    global link_list, table_and_rows, count, filename, whole_page_load, header
    with SB(uc=True, incognito=True, maximize=True, locale_code="en", skip_js_waits=True, headless=False) as sb:
        settings.SWITCH_TO_NEW_TABS_ON_CLICK=True 
        if start_num<=1 and not os.path.exists(f"{final_store_path}/00_part_numbers_links.csv"):
            preparing_link_file(sb.driver, url, final_store_path)
        else:
            print("File with links already exists. Starting Scraping... ")
            print()
        sleep(uniform(sleep_time, sleep_time+2))
        
        #Reading the links from recently made file.
        link_df = pd.read_csv(f"{final_store_path}/00_part_numbers_links.csv")
        link_list = list(link_df['Link Href'])
        
        table_rows_path = final_store_path + "/table_rows_count"
        create_folder(table_rows_path)
        
        for count, link in enumerate(link_list,1):   
            if count < 10:
                count = "0"+str(count)
            ##Start Data Extraction
            filename = link.split("/")[-2]
            filename = clean_file_folder_name(filename)
            logging.info(f'Data Extraction: {filename}')
            file_path = f'{final_store_path}/{count}_{filename}.xlsx'
            sleep(uniform(sleep_time, sleep_time+2))
            create_file(file_path)
            count_file_path = f'{table_rows_path}/{count}_{filename}_count.xlsx'
            
            #Extract count file path if not exits here:
            if not os.path.exists(count_file_path):
                # #Checking if file already exists. It will skip this loop if file exists. If file exist but there is no data inside, we will need to manually correct it.
                sb.driver.get(link)
                sleep(uniform(sleep_time, sleep_time+2))
                #Loading Website on cache
                scroll_whole_page(sb.driver)
                whole_page_load = False
                tables = sb.driver.find_elements(By.XPATH, "//table")
                sleep(uniform(sleep_time, sleep_time+2))
                print("Counting table len and getting part numbers.")
                for table_index in tqdm(range(len(tables)), desc="Processing", leave=True):
                    temp_dict = {}
                    table = tables[table_index]
                    scroll_to_view(sb.driver, table)
                    sleep(uniform(sleep_time, sleep_time+2))
                    table_rows_unfiltered = table.find_elements(By.XPATH, ".//tr")
                    table_rows = []
                    for r in table_rows_unfiltered:
                        try:
                            tr = r.find_element(By.XPATH, ".//a[contains(@class,'PartNbrLnk')]")
                        except:
                            tr = None
                        if tr != None:
                            table_rows.append(r)
                    part_numbers = []
                    
                    for row_index, table_row in enumerate(table_rows,1):
                        table_datas = table_row.find_elements(By.XPATH, ".//td")
                        part_number_link = None
                        for td in table_datas:
                            try:
                                part_number_link = td.find_element(By.XPATH, ".//a")
                                part_number_text = part_number_link.text.strip()  # Get the text and remove leading/trailing spaces
                                if part_number_text and part_number_text[0].isdigit() and part_number_text not in part_numbers:
                                    part_numbers.append(part_number_text)

                             
                            except:
                                continue
                        
                    temp_dict['tables'] = f'table_{table_index+1}'
                    temp_dict[f"number_of_rows"] = len(table_rows)
                    temp_dict["part_numbers"] = part_numbers
                    table_and_rows.append(temp_dict)
                    
                else:
                    # Create a DataFrame from the dictionary
                    df = pd.DataFrame(table_and_rows)
                    # Write the DataFrame to an Excel file
                    
                    
                    print("Writing the tables information to file.")
                    df.to_excel(count_file_path, sheet_name = "RowCount", index=False)
                    table_and_rows.clear()
            
            if os.path.exists(file_path):
                #Loading table and row count file
                count_df = pd.read_excel(count_file_path, sheet_name = "RowCount")
                tables_row_count_list = list(count_df.itertuples(index=False, name=None))
                #Loading Website on cache
                sb.driver.get(link)
                sleep(uniform(sleep_time, sleep_time+2))
                if whole_page_load:
                    scroll_whole_page(sb.driver)
                for sheet_index, table_row_tuple in enumerate(tables_row_count_list,1):
                    #Remaining: Check if all the table/sheet are scraped or not.
                    workbook = openpyxl.load_workbook(file_path)
                    sheet_names = workbook.sheetnames
                    if 'Sheet' in sheet_names:
                        default_sheet = workbook["Sheet"]
                        del workbook[default_sheet.title]

                    # Create a sheet if it doesn't exist
                    if f'Sheet{sheet_index}' not in sheet_names:
                        workbook.create_sheet(f'Sheet{sheet_index}')
                        print(f"Sheet{sheet_index} created.")
                        workbook.save(file_path)
                        sleep(uniform(sleep_time, sleep_time+2))

                    sheet = workbook[f"Sheet{sheet_index}"]
                    first_cell_value = sheet['A1'].value
                    if first_cell_value == 'part_number':
                        print("DataFrame already has a header.")
                    else:
                        # Write headers as the first row
                        headers = ['part_number','brand','retailer','address','pieces_per_package','extracted_date',
                                   'days_to_ship_1',
                                   'quantity_range_start_1','quantity_range_end_1',
                                   'price_1','unit_1',
                                   'days_to_ship_2','quantity_range_start_2','quantity_range_end_2',
                                   'price_2','unit_2'
                        ]
                        # Determine the starting row for writing data
                        start_row = 1

                        # Write the headers to the first row
                        for col, header in enumerate(headers, 1):
                            sheet.cell(row=start_row, column=col, value=header)
                        print("Headers row added to the new sheet.")
                        workbook.save(file_path)
                        sleep(uniform(sleep_time, sleep_time+2))
                    part_numbers_data = ast.literal_eval(table_row_tuple[-1])
                    try:
                        column_data = []
                        for row in sheet.iter_rows(values_only=True):
                            if row[0][0].isdigit():
                                column_data.append(row[0])
                        remaining_to_scrape = []
                        for part_number in part_numbers_data:
                            if part_number not in column_data:
                                remaining_to_scrape.append(part_number)
                    except:
                        remaining_to_scrape = part_numbers_data
                        
                    if remaining_to_scrape == []:
                        remove_duplicates_and_sort(file_path, sheet_index, part_numbers_data)
                        print(f"{table_row_tuple[0]}: All rows are already scraped. Going for next table.")
                        continue
                    else:

                        for remain in remaining_to_scrape:
                            scrape_table_data(sb.driver, sheet_index, remain, final_store_path, file_path)
                        else:
                            print(f"Table {sheet_index}/{table_count} scraped and saved to excel.")
                            print() 

                    remove_duplicates_and_sort(file_path, sheet_index, part_numbers_data)

            else:
                ##Scrape whole page
                pass
        else:
            send_notification(message=f"All links scraped in {pc}.")
            
if __name__ == "__main__":
    for l,p in zip(links,paths): 
        link = base_link + "/" + l
        path = base_path + "/" + p 
        # scrape_data(link,path)
        start_scraping(link, path)
      